"""高效批量文件页数统计工具"""
from pathlib import Path
from time import sleep
from typing import List, Optional, Callable

import pymupdf
import win32com.client as win32
from natsort import os_sorted
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.services.logger_service import setup_logger
from core.document_processing.get_document_statistics import get_document_statistics, WordStatisticType
from core.utils.CountResult import CountResult
from core.utils.open_file_dialog import open_file_dialog
from core.utils.write_report_to_excel import write_report_to_excel

# 设置日志记录器
logger = setup_logger('count_file_pages', 'count_file_pages.log')


class FileType:
    """文件类型分类器"""

    @staticmethod
    def get_type(file_path: Path) -> str:
        ext = file_path.suffix.lower()
        if ext == '.pdf':
            return 'pdf'
        elif ext in ('.doc', '.docx'):
            return 'word'
        elif ext in ('.xls', '.xlsx'):
            return 'excel'
        elif ext in ('.ppt', '.pptx'):
            return 'ppt'
        elif ext in ('.png', '.jpg', '.jpeg'):
            return 'image'
        else:
            return 'unsupported'


class OfficeAppManager:
    """Office 应用程序管理器"""

    def __init__(self):
        self.word_app = None
        self.ppt_app = None

    def get_word_app(self) -> win32.CDispatch:
        """获取或创建 Word 应用实例"""
        if not self.word_app:
            self.word_app = win32.DispatchEx("Word.Application")
            self.word_app.Visible = True
        return self.word_app

    def get_ppt_app(self) -> win32.CDispatch:
        """获取或创建 PowerPoint 应用实例"""
        if not self.ppt_app:
            self.ppt_app = win32.DispatchEx("PowerPoint.Application")
        return self.ppt_app

    def close_all(self):
        """关闭所有 Office 应用程序并释放资源"""
        if self.word_app:
            try:
                # 关闭所有打开的 Word 文档
                while self.word_app.Documents.Count > 0:
                    doc = self.word_app.Documents(1)
                    doc.Close(SaveChanges=False)

                # 退出 Word 应用
                self.word_app.Quit()
            except Exception as e:
                print(f"关闭 Word 失败: {str(e)}")
            finally:
                self.word_app = None

        if self.ppt_app:
            try:
                # 关闭所有打开的 PPT 文档
                while self.ppt_app.Presentations.Count > 0:
                    pres = self.ppt_app.Presentations(1)
                    pres.Close()

                # 退出 PowerPoint 应用
                self.ppt_app.Quit()
            except Exception as e:
                print(f"关闭 PPT 失败: {str(e)}")
            finally:
                self.ppt_app = None


def count_pdf_pages(file_path: Path) -> CountResult:
    """统计 PDF 文件页数"""
    try:
        with pymupdf.open(file_path) as doc:
            logger.info(f"PDF 文件页数统计成功: {file_path.name} - {len(doc)} 页")
            return CountResult(file_path, len(doc))
    except Exception as e:
        logger.error(f"PDF 处理错误: {file_path.name} - {str(e)}")
        return CountResult(file_path, error=f"PDF 处理错误: {str(e)}")


def count_word_pages(word_app: win32.CDispatch, file_path: Path) -> CountResult:
    """使用已打开的 Word 应用统计页数"""
    try:
        doc = word_app.Documents.Open(str(file_path))
        sleep(1) # 等待文件启动
        page_count = get_document_statistics(document=doc, statistic_type=WordStatisticType.PAGES)
        doc.Close(SaveChanges=False)
        logger.info(f"Word 文件页数统计成功: {file_path.name} - {page_count} 页")
        return CountResult(file_path, page_count)
    except Exception as e:
        logger.error(f"Word 处理错误: {file_path.name} - {str(e)}")
        return CountResult(file_path, error=f"Word 处理错误: {str(e)}")


def count_ppt_pages(ppt_app: win32.CDispatch, file_path: Path, include_hidden: bool = False) -> CountResult:
    """使用已打开的 PPT 应用统计页数（默认不包含隐藏页面）"""
    try:
        pres = ppt_app.Presentations.Open(str(file_path), WithWindow=False)

        # 统计可见页数
        visible_slides = 0
        for slide in pres.Slides:
            if include_hidden or not slide.SlideShowTransition.Hidden:
                visible_slides += 1

        pres.Close()
        logger.info(f"PPT 文件页数统计成功: {file_path.name} - {visible_slides} 页")
        return CountResult(file_path, visible_slides)
    except Exception as e:
        logger.error(f"PPT 处理错误: {file_path.name} - {str(e)}")
        return CountResult(file_path, error=f"PPT 处理错误: {str(e)}")


def count_excel_pages(file_path: Path) -> CountResult:
    """Excel 文件不统计页数，返回特殊标记"""
    return CountResult(file_path, page_count=0)


def count_image_pages(file_path: Path) -> CountResult:
    """Image 直接返回 1 页"""
    return CountResult(file_path, page_count=1)


def batch_count_file_pages(
        file_paths: list[Path],
        progress_callback: Callable[..., None] = None
) -> List[CountResult]:
    """高效批量统计文件页数
    
    参数:
        file_paths: 文件路径列表
        progress_callback: 进度更新回调函数
    """
    results: list[CountResult] = []
    app_manager = OfficeAppManager()
    total_files = len(file_paths)
    current_file_num = 0  # 计数
    
    logger.info(f"开始批量统计 {total_files} 个文件的页数")

    try:
        # 按文件类型分组
        file_groups = {}
        for file_path in file_paths:
            file_type = FileType.get_type(file_path)
            if file_type not in file_groups:
                file_groups[file_type] = []
            file_groups[file_type].append(file_path)

        # 处理Word文件
        if 'word' in file_groups:
            word_app = app_manager.get_word_app()
            for i, file_path in enumerate(file_groups['word'], 1):
                current_file_num += 1
                if progress_callback:
                    progress_callback(current_file_num, total_files, f"处理: {file_path.name}")
                results.append(count_word_pages(word_app, file_path))

        # 处理PPT文件
        if 'ppt' in file_groups:
            ppt_app = app_manager.get_ppt_app()
            for i, file_path in enumerate(file_groups['ppt'], 1):
                current_file_num += 1
                if progress_callback:
                    progress_callback(current_file_num, total_files, f"处理: {file_path.name}")
                results.append(count_ppt_pages(ppt_app, file_path))

        # 处理PDF文件
        if 'pdf' in file_groups:
            for i, file_path in enumerate(file_groups['pdf'], 1):
                current_file_num += 1
                if progress_callback:
                    progress_callback(current_file_num, total_files, f"处理: {file_path.name}")
                results.append(count_pdf_pages(file_path))

        # 处理Excel文件
        if 'excel' in file_groups:
            for i, file_path in enumerate(file_groups['excel'], 1):
                current_file_num += 1
                if progress_callback:
                    progress_callback(current_file_num, total_files, f"处理: {file_path.name}")
                results.append(count_excel_pages(file_path))

        if 'image' in file_groups:
            for i, file_path in enumerate(file_groups['image'], 1):
                current_file_num += 1
                if progress_callback:
                    progress_callback(current_file_num, total_files, f"处理: {file_path.name}")
                results.append(count_image_pages(file_path))

        # 处理不支持的文件
        if 'unsupported' in file_groups:
            for i, file_path in enumerate(file_groups['unsupported'], 1):
                current_file_num += 1
                if progress_callback:
                    progress_callback(current_file_num, total_files, f"不支持的文件类型: {file_path.name}")
                results.append(CountResult(file_path, error="不支持的文件类型"))

    except Exception as e:
        # 整体错误处理
        logger.error(f"批量处理失败: {str(e)}", exc_info=True)
        error_result = CountResult(Path(""), error=f"批量处理失败: {str(e)}")
        results = [error_result] * total_files

    finally:
        # 确保释放所有资源
        app_manager.close_all()
        logger.info("所有 Office 应用已关闭")

    return results


def generate_page_count_report(
        results: list[CountResult],
        output_dir: Path
) -> Path:
    """生成页数统计报告"""
    # 处理结果
    report_data = [result.to_row() for result in results]

    # Windows 排序
    report_data = os_sorted(report_data, key=lambda r: r[0])

    # 标题行
    column_header = ["文件名称", "页数"]

    # 生成报告文件路径
    report_path = output_dir / "000--文件页数统计报告.xlsx"

    # 写入文件
    write_report_to_excel(report_data=report_data,
                          column_headers=column_header,
                          output_path=report_path)

    return report_path


def process_page_count_collection(
        progress_callback: Optional[Callable[..., None]] = None
) -> None:
    """带UI的页数统计流程
    
    参数:
        progress_callback: 进度更新回调函数
    """
    try:
        # 选择文件
        if progress_callback:
            progress_callback(message="请选择要统计的文件...")
        logger.info("准备选择文件进行页数统计")

        file_paths = open_file_dialog(
            "选择文件",
            file_filter=[
                ("所有文件", "*.*"),
                ("PDF文件", "*.pdf"),
                ("Word文档", "*.doc*"),
                ("PPT演示文稿", "*.ppt*"),
                ("Excel文件", "*.xls*")
            ],
            multi_select=True
        )

        if not file_paths:
            if progress_callback:
                progress_callback(message="已取消选择文件")
            logger.info("用户取消了文件选择")
            return

        # 转换路径对象
        file_paths = [Path(f) for f in file_paths]
        output_dir = file_paths[0].parent
        logger.info(f"已选择 {len(file_paths)} 个文件，输出目录: {output_dir}")

        # 批量统计
        if progress_callback:
            progress_callback(message=f"开始统计 {len(file_paths)} 个文件的页数...")
        results = batch_count_file_pages(file_paths, progress_callback)

        # 生成报告
        report_path = generate_page_count_report(results, output_dir)
        logger.info(f"页数统计报告已生成: {report_path}")

        # 美化格式
        format_excel_file(str(report_path))

        if progress_callback:
            progress_callback(len(file_paths), len(file_paths), f"统计完成！报告已保存至: {report_path}")

    except Exception as e:
        logger.error(f"页数统计失败: {str(e)}", exc_info=True)
        if progress_callback:
            progress_callback(message=f"统计失败: {str(e)}")


def format_excel_file(file_path):
    """
    使用 Python 的 openpyxl 库，完全替代 VBA 宏 FormatFileTreeSheet。
    执行与 VBA 代码完全相同的格式化操作。

    :param file_path: 需要格式化的 Excel 文件路径（.xlsx 或 .xlsm）
    :raises: 文件操作或格式化的相关异常
    """
    try:
        # 1. 加载工作簿并获取活动工作表
        wb = load_workbook(file_path)
        ws = wb.active

        # 2. 修改整个工作表的字体（对应 VBA 中的 `With Cells.Font`）
        default_font = Font(
            name='等线',         # .Name = "等线"
            size=11,            # .Size = 11
            color='FF000000',   # .ColorIndex = xlAutomatic (黑色)
            bold=False,         # .Bold = False
            italic=False,       # .Italic = False
            underline='none'    # .Underline = False
        )
        # 应用左对齐、顶部对齐
        # left_top_alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')

        # 优化：只遍历有数据的区域
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.font = default_font
                # cell.alignment = left_top_alignment

        # 3. 标题行格式：设置第一行中除A1外的其他单元格居中加粗
        for cell in ws[1]:  # 第 1 行
            cell.alignment = Alignment(horizontal='center', vertical='top')
            cell.font = Font(name='等线', size=11, bold=True)

        # 4. 冻结窗格（对应 ActiveWindow.FreezePanes = True）
        # 冻结第前两行
        ws.freeze_panes = 'A2'

        # 5. 设置列宽（对应 Columns("A:A").ColumnWidth = 50, Columns("B:B").ColumnWidth = 70）
        ws.column_dimensions['A'].width = 70.0
        ws.column_dimensions['B'].width = 12.0

        # 6. 保存更改
        wb.save(file_path)
        logger.info(f"文件格式化完成: {file_path}")

    except Exception as e:
        # 记录错误并重新抛出，方便上层函数处理
        logger.error(f"格式化文件时出错 {file_path}: {e}", exc_info=True)
        raise