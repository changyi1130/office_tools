"""
读取目录树，并写入 Excel，名命为「项目明细表」
"""
from typing import Callable
from pathlib import Path
from natsort import os_sorted
from typing import List
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.utils.run_vba_macro import run_vba_macro
from core.utils.write_report_to_excel import write_report_to_excel
from core.tasks.rename_files import select_directory, is_file_hidden

def read_dirtree(update_info: Callable[[str], None]):
    """ 扫描选定目录，将文件列表生成结构化 Excel 报告 """

    # 选择目录
    selected_dir = select_directory()
    if not selected_dir:
        update_info("操作已取消：未选择目录")
        return

    update_info(f"正在扫描目录，请稍候...")

    # 获取文件列表（原始字符串路径）
    raw_file_paths = get_files(selected_dir)
    if not raw_file_paths:
        update_info(f"目录中没有可处理的文件：{selected_dir}")
        return

    # 转换为 Path 对象以便操作
    file_path_objects = [Path(path) for path in raw_file_paths]

    # 构建数据结构（字典）
    excel_data = { "文件路径" : [], "文件名称" : []}

    for path_obj in file_path_objects:
        excel_data["文件路径"].append(path_obj.parent)
        excel_data["文件名称"].append(path_obj.name)

    # 输出到Excel
    output_excel_path = selected_dir / "000--项目明细表.xlsx"
    write_report_to_excel(report_data=excel_data,
                          output_path=output_excel_path,
                          merge_first_column=True)

    # 美化报告格式
    try:
        # 导入格式化函数，请根据你的项目结构调整导入路径
        # 例如：from core.utils.excel_formatter import format_excel_file
        format_excel_file(str(output_excel_path))
        update_info(f"已将项目明细表存至 {output_excel_path}")
    except Exception as e:
        # 如果格式化失败，报告仍然生成，只是格式未美化
        update_info(f"文件已生成但格式化失败: {e}")

    update_info(f"已将项目明细表存至 {output_excel_path}")


def get_files(path: Path) -> List[str]:
    """获取目录中排序后的文件和子目录"""
    try:
        """获取目录下所有文件（包括子文件）"""
        # 符号链接，避免循环
        if path.is_symlink():
            return []

        items = list(path.iterdir())
        all_files = []

        # 分离文件和目录
        dirs = []
        files = []
        for item in items:
            if item.is_dir():
                dirs.append(item)
            elif item.is_file() and not is_file_hidden(str(item)):
                files.append(str(item))

        # 按 Windows 习惯排序
        files = os_sorted(files)
        dirs = os_sorted(dirs)

        # 添加当前目录的文件
        all_files.extend(files)

        # 递归处理子目录
        for d in dirs:
            all_files.extend(get_files(d))

        return all_files

    except PermissionError as e:
        print(f"目录无权限：{e}")
        return []
    except Exception as e:
        print(f"获取目录内容错误：{e}")
        return []


def format_excel_file(file_path):
    """
    使用 Python 的 openpyxl 库，完全替代 VBA 宏 FormatFileTreeSheet。
    执行与 VBA 代码完全相同的格式化操作。

    :param file_path: 需要格式化的 Excel 文件路径（.xlsx 或 .xlsm）
    :raises: 文件操作或格式化的相关异常
    """
    try:
        # 1. 加载工作簿并获取活动工作表
        wb = load_workbook(file_path)
        ws = wb.active

        # 2. 在 A1 单元格写入标题，并设置格式
        title_cell = ws['A1']
        title_cell.value = '项目明细表'
        ws.merge_cells('A1:B1')
        title_cell.font = Font(name='等线', size=11, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='top')

        # 3. 修改整个工作表的字体（对应 VBA 中的 `With Cells.Font`）
        default_font = Font(
            name='等线',         # .Name = "等线"
            size=11,            # .Size = 11
            color='FF000000',   # .ColorIndex = xlAutomatic (黑色)
            bold=False,         # .Bold = False
            italic=False,       # .Italic = False
            underline='none'    # .Underline = False
        )
        # 应用左对齐、顶部对齐
        left_top_alignment = Alignment(wrap_text=True, horizontal='left', vertical='top')

        # 优化：只遍历有数据的区域
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.font = default_font
                cell.alignment = left_top_alignment

        # 4. 标题行格式：设置第一行中除A1外的其他单元格居中加粗
        for cell in ws[2]:  # 第 2 行
            cell.alignment = Alignment(horizontal='center', vertical='top')
            cell.font = Font(name='等线', size=11, bold=True)

        # 5. 冻结窗格（对应 ActiveWindow.FreezePanes = True）
        # 冻结第前两行
        ws.freeze_panes = 'A3'

        # 6. 设置列宽（对应 Columns("A:A").ColumnWidth = 50, Columns("B:B").ColumnWidth = 70）
        ws.column_dimensions['A'].width = 50.0
        ws.column_dimensions['B'].width = 70.0

        # 7. 保存更改
        wb.save(file_path)
        print(f"✅ 文件格式化完成: {file_path}")

    except Exception as e:
        # 记录错误并重新抛出，方便上层函数处理
        print(f"❌ 格式化文件时出错 {file_path}: {e}")
        raise