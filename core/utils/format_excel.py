"""Excel 文件格式化工具（字体、对齐、冻结窗格、列宽等）"""

from pathlib import Path
from core.services.logger_service import setup_logger

logger = setup_logger('format_excel', 'format_excel.log')


def format_excel_file(file_path: str) -> None:
    """
    使用 openpyxl 格式化 Excel 文件。
    执行与 VBA 代码完全相同的格式化操作。

    :param file_path: 需要格式化的 Excel 文件路径（.xlsx 或 .xlsm）
    :raises: 文件操作或格式化的相关异常
    """
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment
    try:
        # 1. 加载工作簿并获取活动工作表
        wb = load_workbook(file_path)
        ws = wb.active

        # 2. 修改数据行的字体
        default_font = Font(
            name='等线',
            size=11,
            color='FF000000',
            bold=False,
            italic=False,
            underline='none'
        )
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.font = default_font

        # 3. 标题行格式：居中加粗
        for cell in ws[1]:
            cell.alignment = Alignment(horizontal='center', vertical='top')
            cell.font = Font(name='等线', size=11, bold=True)

        # 4. 冻结窗格
        ws.freeze_panes = 'A2'

        # 5. 设置列宽
        ws.column_dimensions['A'].width = 70.0
        ws.column_dimensions['B'].width = 16.0

        # 6. 保存更改
        wb.save(file_path)
        logger.info(f"文件格式化完成: {file_path}")

    except Exception as e:
        logger.error(f"格式化文件时出错 {file_path}: {e}", exc_info=True)
        raise
